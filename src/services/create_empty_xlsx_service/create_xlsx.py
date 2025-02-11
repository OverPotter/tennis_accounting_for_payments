import locale

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from src.constants.colors import HEADER_BACKGROUND_COLOR
from src.constants.column_widths import (
    DATE_OF_PAYMENTS_COLUMN_WIDTH,
    EMPTY_COLUMN_WIDTH,
    FULL_NAME_COLUMN_WIDTH,
    NUMBER_OF_TRAINING_AT_THE_END_OF_MONTH_COLUMN_WIDTH,
    NUMBER_OF_TRAINING_AT_THE_START_OF_MONTH_COLUMN_WIDTH,
    PAID_TRAINING_SESSIONS_COLUMN_WIDTH,
    PAYMENT_AMOUNT_COLUMN_WIDTH,
)
from src.constants.paths import TMP_DIR
from src.constants.xlsx_config import (
    XLSX_FILE_NAME,
    XLSX_LAST_SUB_HEADER,
    XLSX_SUB_HEADERS,
)
from src.services.create_empty_xlsx_service.abc import (
    AbstractCreateEmptyTableService,
)
from src.utils.get_number_of_days_in_month import get_number_of_days_in_month


class CreateEmptyExcelTableService(AbstractCreateEmptyTableService):

    def __init__(self):
        self._bright_green_fill = PatternFill(
            start_color=HEADER_BACKGROUND_COLOR,
            end_color=HEADER_BACKGROUND_COLOR,
            fill_type="solid",
        )
        self._black_font = Font(color="000000")
        self._thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

    # todo: async
    def create_xlsx_table(self) -> str:
        number_days_of_in_month, current_month_name, current_year = (
            get_number_of_days_in_month()
        )
        locale.setlocale(locale.LC_TIME, "ru_RU.UTF-8")

        wb = Workbook()
        ws = wb.active

        self._create_headers(ws, current_month_name, number_days_of_in_month)
        self._create_sub_headers(ws, number_days_of_in_month)
        self._set_column_widths(ws, number_days_of_in_month)

        filename = f"{TMP_DIR}/{XLSX_FILE_NAME}_{current_month_name}_{current_year}.xlsx"
        wb.save(filename)
        return filename

    def _create_headers(self, ws, current_month_name, number_days_of_in_month):
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
        ws.merge_cells(
            start_row=1,
            start_column=7,
            end_row=1,
            end_column=6 + number_days_of_in_month,
        )
        ws.merge_cells(
            start_row=1,
            start_column=6 + number_days_of_in_month + 1,
            end_row=1,
            end_column=6 + number_days_of_in_month + 1,
        )

        self._set_cell(
            ws, row=1, column=1, value=current_month_name.capitalize()
        )
        self._set_cell(ws, row=1, column=7, value="Дата")
        self._set_cell(
            ws,
            row=1,
            column=6 + number_days_of_in_month + 1,
            value=XLSX_LAST_SUB_HEADER,
        )

    def _create_sub_headers(self, ws, number_days_of_in_month):
        for i, sub_header in enumerate(XLSX_SUB_HEADERS):
            self._set_cell(ws, row=2, column=i + 1, value=sub_header)

        for day in range(1, number_days_of_in_month + 1):
            self._set_cell(ws, row=2, column=6 + day, value=day)

    @staticmethod
    def _set_column_widths(ws, number_days_of_in_month):
        ws.column_dimensions["A"].width = FULL_NAME_COLUMN_WIDTH
        ws.column_dimensions["B"].width = (
            NUMBER_OF_TRAINING_AT_THE_START_OF_MONTH_COLUMN_WIDTH
        )
        ws.column_dimensions["C"].width = PAYMENT_AMOUNT_COLUMN_WIDTH
        ws.column_dimensions["D"].width = PAID_TRAINING_SESSIONS_COLUMN_WIDTH
        ws.column_dimensions["E"].width = DATE_OF_PAYMENTS_COLUMN_WIDTH
        ws.column_dimensions["F"].width = EMPTY_COLUMN_WIDTH

        last_column_index = 6 + number_days_of_in_month + 1
        last_column_letter = ws.cell(
            row=1, column=last_column_index
        ).column_letter
        ws.column_dimensions[last_column_letter].width = (
            NUMBER_OF_TRAINING_AT_THE_END_OF_MONTH_COLUMN_WIDTH
        )

    def _set_cell(self, ws, row, column, value):
        cell = ws.cell(row=row, column=column, value=value)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = self._bright_green_fill
        cell.font = self._black_font
        cell.border = self._thin_border
