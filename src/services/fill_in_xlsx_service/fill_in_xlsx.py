from typing import List

from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Border, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet

from src.constants.cell_hieght import CELL_WITH_CLIENTS_DATA_HEIGHT
from src.constants.colors import (
    BORDER_COLOR,
    INCOME_BACKGROUND_COLOR,
    NEGATIVE_TRAINING_COUNT_COLOR,
    NEUTRAL_TRAINING_COUNT_COLOR,
    POSITIVE_TRAINING_COUNT_COLOR,
    TABLE_BACKGROUND_COLOR,
)
from src.constants.offsets import (
    OFFSET_AFTER_HEADERS_FOR_XLSX,
    OFFSET_BETWEEN_CLIENTS,
)
from src.constants.xlsx_config import CELL_WITH_TOTAL_INCOME_NAME
from src.schemas.response.client.monthly_income_and_clients_data import (
    MonthlyIncomeAndClientsDataResponse,
)
from src.schemas.response.month.base import MonthDetailsBaseResponse
from src.schemas.response.payment.base import PaymentBaseResponse
from src.schemas.response.visit.base import VisitBaseResponse
from src.services.fill_in_xlsx_service.abc import AbstractFillInXlsxService
from src.utils.get_training_type_and_number_by_amount import (
    get_training_type_and_number_by_amount,
)
from src.utils.time_helpers.get_current_month_details import (
    get_current_month_details,
)


class FillInXlsxService(AbstractFillInXlsxService):
    def __init__(self) -> None:
        self._current_month_info: MonthDetailsBaseResponse = (
            get_current_month_details()
        )

        self._light_green_fill = PatternFill(
            start_color=POSITIVE_TRAINING_COUNT_COLOR,
            end_color=POSITIVE_TRAINING_COUNT_COLOR,
            fill_type="solid",
        )
        self._light_yellow_fill = PatternFill(
            start_color=NEGATIVE_TRAINING_COUNT_COLOR,
            end_color=NEGATIVE_TRAINING_COUNT_COLOR,
            fill_type="solid",
        )
        self._light_gray_fill = PatternFill(
            start_color=NEUTRAL_TRAINING_COUNT_COLOR,
            end_color=NEUTRAL_TRAINING_COUNT_COLOR,
            fill_type="solid",
        )
        self._light_background_fill = PatternFill(
            start_color=TABLE_BACKGROUND_COLOR,
            end_color=TABLE_BACKGROUND_COLOR,
            fill_type="solid",
        )
        self._income_background_fill = PatternFill(
            start_color=INCOME_BACKGROUND_COLOR,
            end_color=INCOME_BACKGROUND_COLOR,
            fill_type="solid",
        )

        self._bold_side = Side(border_style="thick", color=BORDER_COLOR)
        self._thin_side = Side(border_style="thin", color=BORDER_COLOR)

    def fill_in_xlsx(
        self,
        total_income_and_clients_data: MonthlyIncomeAndClientsDataResponse,
        filename: str,
    ) -> None:
        wb = load_workbook(filename)
        ws = wb.active
        row_after_headers = OFFSET_AFTER_HEADERS_FOR_XLSX

        for client in total_income_and_clients_data.clients_data:
            client_start = row_after_headers

            self._fill_client_name(ws, client_start, client.name)
            self._fill_visits_beginning(
                ws, client_start, client.visits_at_the_beginning_of_the_month
            )
            self._fill_visits(ws, client_start, client.monthly_visits)
            self._fill_visits_end(
                ws, client_start, client.visits_at_the_end_of_the_month
            )

            payments_offset = self._fill_payments(
                ws, client_start, client.monthly_payments
            )
            if payments_offset > 0:
                block_end = client_start + payments_offset - 1
            else:
                block_end = client_start

            self._apply_block_format(ws, client_start, block_end)

            row_after_headers = block_end + OFFSET_BETWEEN_CLIENTS + 1

        self._add_row_with_total_income(
            ws=ws,
            row_after_headers=row_after_headers,
            total_income=total_income_and_clients_data.total_income,
        )

        wb.save(filename)

    @staticmethod
    def _fill_client_name(ws: Worksheet, row: int, name: str) -> None:
        ws.cell(row=row, column=1, value=name)

    def _fill_visits_beginning(
        self, ws: Worksheet, row: int, visits: int
    ) -> None:
        cell = ws.cell(row=row, column=2, value=visits)
        cell.fill = self._get_fill_color(visits)

    @staticmethod
    def _fill_visits(
        ws: Worksheet, row: int, visits: List[VisitBaseResponse]
    ) -> None:
        visits_by_day = {}
        for visit in visits:
            day = visit.visit_datetime.day
            visits_by_day[day] = visits_by_day.get(day, 0) + 1
        for day, count in visits_by_day.items():
            ws.cell(row=row, column=6 + day, value=count)

    def _fill_visits_end(self, ws: Worksheet, row: int, visits: int) -> None:
        last_col = 6 + self._current_month_info.total_days + 1
        cell = ws.cell(row=row, column=last_col, value=visits)
        cell.fill = self._get_fill_color(visits)

    @staticmethod
    def _fill_payments(
        ws: Worksheet, row: int, payments: List[PaymentBaseResponse]
    ) -> int:
        offset = 0
        for payment in payments:
            ws.cell(row=row + offset, column=3, value=payment.amount)
            number_of_paid_training = get_training_type_and_number_by_amount(
                payment.amount
            )[0]
            ws.cell(row=row + offset, column=4, value=number_of_paid_training)
            ws.cell(
                row=row + offset,
                column=5,
                value=payment.payment_date.strftime("%d.%m.%Y"),
            )
            offset += 1
        return offset

    def _get_fill_color(self, visits_count: int) -> PatternFill:
        if visits_count > 0:
            return self._light_green_fill
        elif visits_count < 0:
            return self._light_yellow_fill
        return self._light_gray_fill

    def _apply_block_format(
        self, ws: Worksheet, block_start: int, block_end: int
    ) -> None:
        last_col = 6 + self._current_month_info.total_days + 1

        for r in range(block_start, block_end + 1):
            ws.row_dimensions[r].height = CELL_WITH_CLIENTS_DATA_HEIGHT

            for c in range(1, last_col + 1):
                cell = ws.cell(row=r, column=c)

                if cell.fill is None or cell.fill.patternType is None:
                    cell.fill = self._light_background_fill

                top_side = (
                    self._bold_side if r == block_start else self._thin_side
                )
                bottom_side = (
                    self._bold_side if r == block_end else self._thin_side
                )
                left_side = self._bold_side if c == 1 else self._thin_side
                right_side = (
                    self._bold_side if c == last_col else self._thin_side
                )

                cell.border = Border(
                    top=top_side,
                    bottom=bottom_side,
                    left=left_side,
                    right=right_side,
                )

        self._fill_in_the_offset_lines_with_color(
            ws=ws, last_col=last_col, block_end=block_end
        )

    def _fill_in_the_offset_lines_with_color(
        self, ws: Worksheet, last_col: int, block_end: int
    ):
        if OFFSET_BETWEEN_CLIENTS > 0:
            empty_row = block_end + 1
            ws.row_dimensions[empty_row].height = 15
            for c in range(1, last_col + 1):
                ws.cell(row=empty_row, column=c).fill = (
                    self._light_background_fill
                )

    def _add_row_with_total_income(
        self, ws: Worksheet, row_after_headers: int, total_income: float
    ) -> None:
        row_for_total = row_after_headers
        ws.cell(row=row_for_total, column=1, value=CELL_WITH_TOTAL_INCOME_NAME)
        ws.cell(row=row_for_total, column=2, value="")
        ws.cell(row=row_for_total, column=3, value=total_income)

        self._apply_border_to_total_income_row(
            ws=ws, row_for_total=row_for_total
        )

        for col in range(1, 4):
            cell = ws.cell(row=row_for_total, column=col)
            cell.fill = self._income_background_fill

    def _apply_border_to_total_income_row(
        self, ws: Worksheet, row_for_total: int
    ) -> None:

        for col in range(1, 4):
            cell = ws.cell(row=row_for_total, column=col)

            cell.border = Border(
                top=self._bold_side if col == 1 else None,
                bottom=self._bold_side,
                left=self._bold_side if col == 1 else None,
                right=self._bold_side if col == 3 else None,
            )
