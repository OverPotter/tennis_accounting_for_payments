import asyncio
import locale

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from src.constants import TMP_DIR
from src.services.create_xlsx.abc import AbstractCreateTableService
from src.utils.get_number_of_days_in_month import get_number_of_days_in_month


class CreateExcelTableService(AbstractCreateTableService):
    @staticmethod
    async def create_xlsx_table():
        number_days_of_in_month, current_month_name, current_year = (
            get_number_of_days_in_month()
        )

        locale.setlocale(locale.LC_TIME, "ru_RU.UTF-8")

        wb = Workbook()
        ws = wb.active

        sub_headers_month = [
            "ФИО",
            "Остаточный переход",
            "Количество",
            "Сумма оплаты",
            "Количество тренировок",
            "Дата оплаты",
        ]

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
        ws.merge_cells(
            start_row=1,
            start_column=7,
            end_row=1,
            end_column=6 + number_days_of_in_month,
        )
        ws.merge_cells(
            start_row=1,
            start_column=6 + number_days_of_in_month + 1,
            end_row=1,
            end_column=6 + number_days_of_in_month + 1,
        )

        bright_green_fill = PatternFill(
            start_color="9ACD32", end_color="9ACD32", fill_type="solid"
        )
        black_font = Font(color="000000")

        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        ws.cell(row=1, column=1, value=current_month_name.capitalize()).alignment = (
            Alignment(horizontal="center", vertical="center")
        )
        ws.cell(row=1, column=1).fill = bright_green_fill
        ws.cell(row=1, column=1).font = black_font
        ws.cell(row=1, column=1).border = thin_border

        ws.cell(row=1, column=7, value="Дата").alignment = Alignment(
            horizontal="center", vertical="center"
        )
        ws.cell(row=1, column=7).fill = bright_green_fill
        ws.cell(row=1, column=7).font = black_font
        ws.cell(row=1, column=7).border = thin_border

        ws.cell(
            row=1,
            column=6 + number_days_of_in_month + 1,
            value="Количество оставшихся тренировок",
        ).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=1, column=6 + number_days_of_in_month + 1).fill = bright_green_fill
        ws.cell(row=1, column=6 + number_days_of_in_month + 1).font = black_font
        ws.cell(row=1, column=6 + number_days_of_in_month + 1).border = thin_border

        for col in range(1, 7 + number_days_of_in_month + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = bright_green_fill
            cell.font = black_font
            cell.border = thin_border

        for i, sub_header in enumerate(sub_headers_month):
            cell = ws.cell(row=2, column=i + 1, value=sub_header)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = bright_green_fill
            cell.font = black_font
            cell.border = thin_border

        for day in range(1, number_days_of_in_month + 1):
            cell = ws.cell(row=2, column=6 + day, value=day)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = bright_green_fill
            cell.font = black_font
            cell.border = thin_border

        ws.column_dimensions["B"].width = 20  # Остаточный переход
        ws.column_dimensions["C"].width = 15  # Количество
        ws.column_dimensions["D"].width = 20  # Сумма оплаты
        ws.column_dimensions["E"].width = 20  # Количество тренировок
        ws.column_dimensions["F"].width = 15  # Дата оплаты

        last_column_index = 6 + number_days_of_in_month + 1
        last_column_letter = ws.cell(row=1, column=last_column_index).column_letter
        ws.column_dimensions[last_column_letter].width = 35

        ws.oddHeader.left.text = "left"
        ws.oddHeader.center.text = "center"
        ws.oddHeader.right.text = "right"
        filename = f"{TMP_DIR}/{current_month_name}_{current_year}.xlsx"
        wb.save(filename)


if __name__ == "__main__":
    asyncio.run(CreateExcelTableService.create_xlsx_table())
