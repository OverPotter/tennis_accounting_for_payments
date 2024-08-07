import locale

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from src.constants import TMP_DIR
from src.services.create_xlsx_service.abc import AbstractCreateTableService
from src.utils.get_number_of_days_in_month import get_number_of_days_in_month


class CreateExcelTableService(AbstractCreateTableService):

    def __init__(self):
        self._bright_green_fill = PatternFill(
            start_color="9ACD32", end_color="9ACD32", fill_type="solid"
        )
        self._black_font = Font(color="000000")
        self._thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

    async def create_xlsx_table(self):
        number_days_of_in_month, current_month_name, current_year = (
            get_number_of_days_in_month()
        )
        locale.setlocale(locale.LC_TIME, "ru_RU.UTF-8")

        wb = Workbook()
        ws = wb.active

        self._create_headers(ws, current_month_name, number_days_of_in_month)
        self._create_sub_headers(ws, number_days_of_in_month)
        self._set_column_widths(ws, number_days_of_in_month)

        filename = f"{TMP_DIR}/{current_month_name}_{current_year}.xlsx"
        wb.save(filename)

    def _create_headers(self, ws, current_month_name, number_days_of_in_month):
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
        ws.merge_cells(
            start_row=1,
            start_column=7,
            end_row=1,
            end_column=6 + number_days_of_in_month,
        )
        ws.merge_cells(
            start_row=1,
            start_column=6 + number_days_of_in_month + 1,
            end_row=1,
            end_column=6 + number_days_of_in_month + 1,
        )

        self._set_cell(
            ws, row=1, column=1, value=current_month_name.capitalize()
        )
        self._set_cell(ws, row=1, column=7, value="Дата")
        self._set_cell(
            ws,
            row=1,
            column=6 + number_days_of_in_month + 1,
            value="Количество оставшихся тренировок",
        )

    def _create_sub_headers(self, ws, number_days_of_in_month):
        sub_headers_month = [
            "ФИО",
            "Остаточный переход",
            "Количество",
            "Сумма оплаты",
            "Количество тренировок",
            "Дата оплаты",
        ]

        for i, sub_header in enumerate(sub_headers_month):
            self._set_cell(ws, row=2, column=i + 1, value=sub_header)

        for day in range(1, number_days_of_in_month + 1):
            self._set_cell(ws, row=2, column=6 + day, value=day)

    def _set_column_widths(self, ws, number_days_of_in_month):
        ws.column_dimensions["B"].width = 20  # Transition
        ws.column_dimensions["C"].width = 15  # Quantity
        ws.column_dimensions["D"].width = 20  # Payment amount
        ws.column_dimensions["E"].width = 20  # Number of training sessions
        ws.column_dimensions["F"].width = 15  # Date of payment

        last_column_index = 6 + number_days_of_in_month + 1
        last_column_letter = ws.cell(
            row=1, column=last_column_index
        ).column_letter
        ws.column_dimensions[last_column_letter].width = 35

    def _set_cell(self, ws, row, column, value):
        cell = ws.cell(row=row, column=column, value=value)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = self._bright_green_fill
        cell.font = self._black_font
        cell.border = self._thin_border
